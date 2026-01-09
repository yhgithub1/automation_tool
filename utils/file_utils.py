import os
import glob
from typing import Tuple, Optional, List, Any
import logging

# Lazy import for openpyxl
def get_openpyxl():
    import openpyxl
    return openpyxl

# 设置日志
logger = logging.getLogger(__name__)


def find_excel_file() -> Tuple[Optional[str], str]:
    """
    在桌面的tool文件夹中查找名为'datasource'的Excel文件

    Returns:
        Tuple[可选的文件路径, 状态消息]
    """
    try:
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        tool_folder = os.path.join(desktop_path, "tool")

        if not os.path.exists(tool_folder):
            return None, "错误: 桌面的tool文件夹不存在"

        # 查找名为'datasource'的Excel文件（不区分大小写）
        excel_files = []
        for pattern in ["*.xlsx", "*.xls"]:
            for file_path in glob.glob(os.path.join(tool_folder, pattern)):
                file_name = os.path.basename(file_path).lower()
                if "datasource" in file_name:
                    excel_files.append(file_path)

        if not excel_files:
            return None, "错误: tool文件夹中未找到名为'datasource'的Excel文件"

        # 使用第一个找到的'datasource' Excel文件
        excel_file = excel_files[0]
        message = f"找到Excel文件: {os.path.basename(excel_file)}"

        # 如果有多个'datasource' Excel文件，添加提示
        if len(excel_files) > 1:
            message += f"\n注意: tool文件夹中发现{len(excel_files)}个名为'datasource'的Excel文件，将使用第一个文件"

        return excel_file, message

    except Exception as e:
        error_msg = f"查找Excel文件时出错: {str(e)}"
        logger.error(error_msg)
        return None, error_msg


def read_excel_data(file_path: str, sheet_name: Optional[str] = None, header_row: Optional[int] = None) -> Optional[List[List[Any]]]:
    """
    读取Excel文件数据

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，如果为None则读取第一个工作表
        header_row: 表头所在行，如果为None则自动检测

    Returns:
        List of lists representing Excel data, or None if error
    """
    try:
        # Load workbook
        openpyxl = get_openpyxl()
        workbook = openpyxl.load_workbook(file_path, read_only=True)

        # Get sheet
        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # Read data
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # Handle header row if specified
        if header_row is not None and header_row > 0 and len(data) > header_row - 1:
            headers = data[header_row - 1]
            data = data[header_row:]
            # Convert to dictionary format if headers exist
            if headers:
                result = []
                for row in data:
                    if len(row) == len(headers):
                        result.append(dict(zip(headers, row)))
                    else:
                        result.append(row)
                return result
            else:
                return data
        else:
            return data

    except Exception as e:
        error_msg = f"读取Excel文件失败: {str(e)}"
        logger.error(error_msg)
        return None
    finally:
        if 'workbook' in locals():
            workbook.close()


def validate_excel_file(file_path: str) -> Tuple[bool, str]:
    """
    验证Excel文件是否有效

    Args:
        file_path: Excel文件路径

    Returns:
        (是否有效, 错误消息)
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return False, "文件不存在"

        # 检查文件扩展名
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            return False, "文件不是有效的Excel格式"

        # 尝试读取文件前几行
        openpyxl = get_openpyxl()
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        # Try to read first row
        next(sheet.iter_rows(max_row=1))
        workbook.close()

        return True, "文件有效"

    except Exception as e:
        return False, f"Excel文件无效: {str(e)}"


def get_sheet_names(file_path: str) -> Optional[list]:
    """
    获取Excel文件中的所有工作表名称

    Args:
        file_path: Excel文件路径

    Returns:
        工作表名称列表或None（如果出错）
    """
    try:
        # 使用openpyxl获取工作表名称
        openpyxl = get_openpyxl()
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()

        return sheet_names

    except Exception as e:
        logger.error(f"获取工作表名称失败: {str(e)}")
        return None


# 简单的测试代码
if __name__ == "__main__":
    # 设置日志
    logging.basicConfig(level=logging.INFO)

    # 测试查找Excel文件
    excel_path, message = find_excel_file()
    print(message)

    if excel_path:
        # 测试验证文件
        is_valid, valid_msg = validate_excel_file(excel_path)
        print(f"文件验证: {is_valid}, {valid_msg}")

        # 测试读取Excel数据
        data = read_excel_data(excel_path)
        if data is not None:
            print(f"成功读取 {len(data)} 行数据")
            print("前5行数据:")
            for i, row in enumerate(data[:5]):
                print(f"行 {i+1}: {row}")

        # 测试获取工作表名称
        sheets = get_sheet_names(excel_path)
        if sheets:
            print(f"工作表名称: {sheets}")
