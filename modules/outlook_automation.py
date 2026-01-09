import win32com.client as win32
import re
import os
import time
import glob
import pythoncom
from PyQt5.QtCore import QThread, pyqtSignal
import sys
import os
from typing import List, Any
import threading

# Lazy import for openpyxl
def get_openpyxl():
    import openpyxl
    return openpyxl

# Get the directory of the current script
current_dir = os.path.dirname(os.path.abspath(__file__))
# Get the project root (parent of modules directory)
project_root = os.path.dirname(current_dir)
# Add project root to Python path if not already there
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from utils.file_utils import find_excel_file

def get_email_addresses_from_datasource():
    """
    从桌面tool文件夹中的datasource.xlsx文件读取邮箱地址
    返回TO和CC邮箱地址列表
    """
    try:
        # 构建datasource.xlsx文件的完整路径
        datasource_path = os.path.join(os.path.expanduser("~"), "Desktop", "tool", "datasource.xlsx")

        # 检查文件是否存在
        if not os.path.exists(datasource_path):
            raise FileNotFoundError(f"datasource.xlsx文件不存在: {datasource_path}")

        # 读取Excel文件中的outlook_tocc表
        openpyxl = get_openpyxl()
        workbook = openpyxl.load_workbook(datasource_path, read_only=True)
        sheet = workbook['outlook_tocc']

        # 初始化邮箱地址字典
        email_addresses = {
            'to': [],
            'cc': []
        }

        # 读取所有数据
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        # 解析表格数据 - 处理新的Excel结构
        # 新结构：列标题包含"to"和TO邮箱地址，数据行包含"cc"和CC邮箱地址
        for row_idx, row in enumerate(data):
            if len(row) == 0:
                continue

            row_type = str(row[0]).lower().strip()  # 第一列是类型（to/cc）

            # 遍历剩余列，提取非空的邮箱地址
            for col_idx in range(1, len(row)):
                email = str(row[col_idx]).strip() if row[col_idx] is not None else ''
                # 更严格的NaN检查
                if email and email.lower() != 'nan' and email != '':
                    if row_type == 'to':
                        email_addresses['to'].append(email)
                    elif row_type == 'cc':
                        email_addresses['cc'].append(email)

        # 如果没有找到TO邮箱地址，尝试从列标题中提取（新Excel结构）
        if not email_addresses['to']:
            # 检查列标题是否包含TO邮箱地址
            if len(data) > 0:
                header_row = data[0]
                for col_idx in range(1, len(header_row)):
                    col_name = str(header_row[col_idx]).strip() if header_row[col_idx] is not None else ''
                    if col_name and col_name.lower() != 'nan' and col_name != '' and '@' in col_name:
                        email_addresses['to'].append(col_name)

        # 如果没有找到TO邮箱地址，使用默认值
        if not email_addresses['to']:
            email_addresses['to'] = ["zicheng.zhang@zeiss.com", "jiaxin.lu.ext@zeiss.com"]

        # 如果没有找到CC邮箱地址，使用默认值
        if not email_addresses['cc']:
            email_addresses['cc'] = ["Zhu, Zhiming"]

        # 将邮箱地址列表转换为Outlook格式（用分号分隔）
        to_emails = ';'.join(email_addresses['to'])
        cc_emails = ';'.join(email_addresses['cc'])

        return to_emails, cc_emails

    except Exception as e:
        # 如果读取失败，返回默认值（与原始硬编码一致）
        print(f"读取datasource.xlsx文件时出错: {str(e)}")
        return "zicheng.zhang@zeiss.com;jiaxin.lu.ext@zeiss.com", "Zhu, Zhiming"


class OutlookEmailThread(QThread):
    """Outlook邮件生成线程（支持COM初始化，处理Excel并生成邮件）"""
    progress = pyqtSignal(str)
    finished = pyqtSignal(bool)

    # 类级别的锁，确保只有一个Outlook线程可以运行
    outlook_lock = threading.Lock()
    outlook_active = False

    def __init__(self, excel_path):
        super().__init__()
        self.excel_path = excel_path

    def run(self):
        # 使用锁机制确保只有一个Outlook线程可以运行
        with OutlookEmailThread.outlook_lock:
            # 检查是否已经有Outlook线程在运行
            if OutlookEmailThread.outlook_active:
                self.progress.emit("⚠️  警告：已有Outlook任务正在运行，请等待完成后重试")
                self.finished.emit(False)
                return

            # 标记Outlook任务开始
            OutlookEmailThread.outlook_active = True

        try:
            # 初始化 COM 环境（必须在操作 Outlook 前调用）
            pythoncom.CoInitialize()
            try:
                result = self._generate_emails_from_excel()
                self.finished.emit(result)
            except Exception as e:
                # 处理 COM 注册问题
                if "CLSIDToClassMap" in str(e):
                    self.progress.emit(f"Outlook COM 组件需要重新注册，尝试修复...")
                    try:
                        # 尝试重新初始化
                        pythoncom.CoUninitialize()
                        pythoncom.CoInitialize()
                        result = self._generate_emails_from_excel()
                        self.finished.emit(result)
                        return
                    except Exception as retry_e:
                        self.progress.emit(f"重新尝试失败: {str(retry_e)}")

                self.progress.emit(f"全局错误: {str(e)}")
                self.finished.emit(False)
            finally:
                # 释放 COM 资源，避免内存泄漏
                pythoncom.CoUninitialize()
        except Exception as e:
            self.progress.emit(f"线程运行时发生意外错误: {str(e)}")
            self.finished.emit(False)
        finally:
            # 任务完成，释放锁
            with OutlookEmailThread.outlook_lock:
                OutlookEmailThread.outlook_active = False
                self.progress.emit("Outlook任务完成，锁已释放")

    def _generate_emails_from_excel(self):
        """核心逻辑：从Excel读取数据，生成Outlook邮件"""
        try:
            # -------- 1. 读取Excel文件 --------
            self.progress.emit("正在读取Excel文件...")
            openpyxl = get_openpyxl()
            workbook = openpyxl.load_workbook(self.excel_path, read_only=True)
            sheet = workbook['Sheet1']

            # 读取所有数据并过滤空行
            data = []
            for row in sheet.iter_rows(values_only=True):
                row_list = list(row)
                # 过滤空行：如果行中所有单元格都是None或空字符串，则跳过
                if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row_list):
                    continue
                data.append(row_list)

            self.progress.emit(f"成功读取Excel文件，共{len(data)}行数据")

            # -------- 2. 启动Outlook --------
            self.progress.emit("启动Outlook应用程序...")
            outlook = self._get_outlook_application()
            if not outlook:
                self.progress.emit("无法启动Outlook应用程序，请确保Outlook已安装且正常运行")
                return False
            self.progress.emit("Outlook应用程序已启动")

            # 检查是否已经有Outlook窗口打开
            try:
                import psutil
                outlook_windows = []
                for proc in psutil.process_iter(['name', 'pid']):
                    if proc.info['name'] and 'outlook' in proc.info['name'].lower():
                        outlook_windows.append(proc.info['pid'])

                if len(outlook_windows) > 1:  # 当前进程 + 现有Outlook
                    self.progress.emit(f"⚠️  检测到已有Outlook窗口打开，将使用现有Outlook实例")
            except ImportError:
                # psutil未安装，跳过检查
                pass
            except Exception as e:
                self.progress.emit(f"检查现有Outlook窗口时出错: {str(e)}")

            # -------- 3. 捕获Outlook签名 --------
            self.progress.emit("正在获取Outlook签名...")
            signature = self._capture_outlook_signature(outlook)
            if not signature or len(signature) <= 50:
                self.progress.emit("警告: 未捕获到有效签名，邮件将不含签名")

            # -------- 4. 循环生成邮件 --------
            total_rows = len(data)
            for index, row in enumerate(data):
                self.progress.emit(f"正在处理第{index + 1}/{total_rows}行数据...")
                try:
                    # 提取公司名称（F列，索引2）
                    company_full = str(row[2]) if len(row) > 2 else ""
                    company_name = company_full.split('/')[-1].strip() if '/' in company_full else company_full.strip()

                    # 提取设备信息（型号：H列索引4；SN：B列索引1）
                    model = str(row[4]) if len(row) > 4 else ""
                    sn = str(row[1]) if len(row) > 1 else ""

                    # 提取地址和联系人（N列，索引13）
                    contact_info = re.sub(r'\s+', ' ', str(row[13])).strip() if len(row) > 13 else ""

                    # 构建邮件主题和正文
                    subject = f"{company_name} {model} SN:{sn}包装箱回收"
                    body = f"""Dear Mr. Zhang:

如下客户请回收包装箱，麻烦尽快安排：
客户：{company_name}
CMM: {model} SN: {sn}
地址及联系人：{contact_info}
谢谢！

"""

                    # 创建HTML格式邮件
                    html_body = f"""
<html>
<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
<style>
    body {{ font-family: 'Calibri', sans-serif; font-size: 11pt; }}
    pre {{ font-family: 'Calibri', sans-serif; font-size: 11pt; }}
</style>
</head>
<body>
<pre>{body}</pre>
<br>
{signature if signature else ''}
</body>
</html>
"""

                    # 创建并显示邮件
                    mail = outlook.CreateItem(0)
                    mail.Subject = subject
                    mail.HTMLBody = html_body

                    # 从datasource.xlsx获取邮箱地址
                    to_emails, cc_emails = get_email_addresses_from_datasource()
                    self.progress.emit(f"使用邮箱地址 - TO: {to_emails}, CC: {cc_emails}")

                    mail.To = to_emails
                    mail.CC = cc_emails
                    mail.Display()

                    self.progress.emit(f"已创建邮件 #{index + 1}: {subject}")
                    time.sleep(1)  # 给Outlook留处理时间

                except Exception as e:
                    self.progress.emit(f"处理行{index + 1}时出错: {str(e)}")

            self.progress.emit("邮件创建完成！请检查Outlook窗口")
            return True

        except Exception as e:
            self.progress.emit(f"生成邮件时出错: {str(e)}")
            return False

    def _capture_outlook_signature(self, outlook):
        """多方法捕获Outlook签名"""
        try:
            # 方法1：从默认路径读取签名文件
            appdata = os.getenv('APPDATA')
            signature_path = os.path.join(appdata, 'Microsoft', 'Signatures')
            if os.path.exists(signature_path):
                html_files = glob.glob(os.path.join(signature_path, '*.htm')) or glob.glob(os.path.join(signature_path, '*.html'))
                if html_files:
                    latest_file = max(html_files, key=os.path.getmtime)
                    with open(latest_file, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    # 修正签名中的图片路径
                    base_name = os.path.splitext(os.path.basename(latest_file))[0]
                    image_folder = os.path.join(signature_path, base_name)
                    if os.path.exists(image_folder):
                        content = content.replace(f'"{base_name}/', f'"{image_folder}/')
                        content = content.replace(f"'{base_name}/", f"'{image_folder}/")
                    return content if len(content) > 50 else ""

            # 方法2：通过临时邮件获取签名
            temp_mail = outlook.CreateItem(0)
            temp_mail.Display()
            time.sleep(2)
            signature = ""
            for _ in range(5):
                signature = temp_mail.HTMLBody
                if len(signature) > 100 and ("<p>" in signature or "<div>" in signature):
                    break
                time.sleep(1)
            temp_mail.Close(0)  # 不保存临时邮件

            # 提取签名（通过<hr>标签分割）
            if "<hr" in signature:
                signature = signature.split("<hr", 1)[-1].split(">", 1)[-1]
            return signature if len(signature) > 50 else ""

        except Exception as e:
            self.progress.emit(f"捕获签名失败: {str(e)}")
            return ""

    def _get_outlook_application(self):
        """获取Outlook应用程序对象，包含多种COM初始化方法"""
        try:
            # 方法1：直接使用Dispatch
            try:
                outlook = win32.Dispatch('Outlook.Application')
                return outlook
            except Exception as e1:
                self.progress.emit(f"直接Dispatch失败: {str(e1)}")
            
            # 方法2：使用gencache确保分发
            try:
                import win32com.client.gencache
                outlook = win32com.client.gencache.EnsureDispatch('Outlook.Application')
                self.progress.emit("使用gencache方法成功")
                return outlook
            except Exception as e2:
                self.progress.emit(f"gencache方法失败: {str(e2)}")
            
            # 方法3：清除并重建COM缓存
            try:
                self.progress.emit("尝试清除损坏的COM缓存...")
                self._clear_com_cache()
                import win32com.client.gencache
                outlook = win32com.client.gencache.EnsureDispatch('Outlook.Application')
                self.progress.emit("清除缓存后成功")
                return outlook
            except Exception as e3:
                self.progress.emit(f"清除缓存后仍失败: {str(e3)}")
            
            # 方法4：检查Outlook是否正在运行并启动
            try:
                import psutil
                outlook_running = any(proc.name().lower() == 'outlook.exe' for proc in psutil.process_iter(['name']))
                if not outlook_running:
                    self.progress.emit("检测到Outlook未运行，尝试启动...")
                    import subprocess
                    subprocess.Popen(['outlook.exe'])
                    time.sleep(10)  # 增加等待时间
                    # 重试
                    outlook = win32.Dispatch('Outlook.Application')
                    return outlook
            except ImportError:
                self.progress.emit("无法检查Outlook进程状态（psutil未安装）")
            except Exception as e4:
                self.progress.emit(f"启动Outlook失败: {str(e4)}")
            
            # 方法5：使用CLSID直接访问
            try:
                outlook = win32.Dispatch('{0006F03A-0000-0000-C000-000000000046}')
                self.progress.emit("使用CLSID方法成功")
                return outlook
            except Exception as e5:
                self.progress.emit(f"CLSID方法也失败: {str(e5)}")
            
            # 如果所有方法都失败，提供详细指导
            self.progress.emit("Outlook连接失败，请按以下步骤手动修复：")
            self.progress.emit("1. 关闭所有Outlook窗口")
            self.progress.emit("2. 按 Win+R，输入 'cmd'，运行以下命令：")
            self.progress.emit("   cd %APPDATA%\\Python\\Pythonwin32\\gen_py")
            self.progress.emit("   rmdir /s /q *")
            self.progress.emit("3. 重新启动Outlook应用程序")
            self.progress.emit("4. 重新运行此工具")
            self.progress.emit("或者尝试：控制面板 → 程序和功能 → 修复Microsoft Office")
            return None

        except Exception as e:
            self.progress.emit(f"获取Outlook应用程序时发生未知错误: {str(e)}")
            return None

    def _clear_com_cache(self):
        """清除损坏的COM缓存文件"""
        try:
            import shutil
            # 获取Python COM缓存目录
            gen_py_dir = os.path.join(os.path.expanduser("~"), "AppData", "Roaming", "Python", "Pythonwin32", "gen_py")
            
            if os.path.exists(gen_py_dir):
                self.progress.emit(f"正在清除COM缓存目录: {gen_py_dir}")
                # 备份当前目录名并删除
                backup_dir = gen_py_dir + "_backup_" + str(int(time.time()))
                if os.path.exists(backup_dir):
                    shutil.rmtree(backup_dir)
                os.rename(gen_py_dir, backup_dir)
                self.progress.emit(f"已备份旧缓存到: {backup_dir}")
            else:
                self.progress.emit("COM缓存目录不存在，无需清除")
                
        except Exception as e:
            self.progress.emit(f"清除COM缓存时出错: {str(e)}")
            # 即使清除失败，继续尝试其他方法


# 测试代码（直接运行该脚本时执行）
if __name__ == "__main__":
    excel_path, msg = find_excel_file()
    if not excel_path:
        print(msg)
        exit(1)
    print(f"找到Excel文件: {excel_path}")
    # 创建并启动线程
    thread = OutlookEmailThread(excel_path)
    thread.progress.connect(print)
    thread.finished.connect(lambda success: print(f"执行完成: {success}"))
    thread.start()
    thread.wait()  # 等待线程结束（仅测试用）
